# -*- coding: utf-8 -*-
"""
숙박업 원장(Dataset B) Import Pipeline
  숙박업 리스트_260220.xlsx → 정규화 → 교차검증 → index.html 에 데이터 블록 주입

원칙 (2026-08-28 지시):
 - Dataset A(등급유효 835)와 Dataset B(숙박업 인허가 원장)는 모집단이 다른 별도 Source.
   어느 한쪽으로 덮어쓰지 않고 교차검증만 한다.
 - Dataset B 의 Primary Key 는 관리번호. 사업장명으로 중복 제거하지 않는다
   (같은 건물에 운영사별 인허가 여러 개 존재 — 예: 엘시티 레지던스 4건).
 - 시트명은 신뢰하지 않는다. 세종(1) 시트는 울산 데이터의 복사본이었음 →
   지역은 반드시 주소에서 파싱하고, 시트 간 중복은 관리번호로 제거한다.
 - 객실수 0/공란은 "객실 없음"이 아니라 "미기재"로 취급한다.
 - 원본 Excel 은 수정하지 않는다.

실행:  python registry_build.py
출력:  data/registry_busan.json  data/registry_summary.json  data/xmatch.json
       + index.html 의 /*==REGISTRY_DATA==*/ 블록 갱신
"""
import openpyxl, sys, json, re, io, os
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, '숙박업 리스트_260220.xlsx')
HTML = os.path.join(BASE, 'index.html')
os.makedirs(os.path.join(BASE, 'data'), exist_ok=True)

SIDO = {
    '서울특별시': '서울', '부산광역시': '부산', '대구광역시': '대구', '인천광역시': '인천',
    '광주광역시': '광주', '대전광역시': '대전', '울산광역시': '울산', '세종특별자치시': '세종',
    '경기도': '경기', '강원특별자치도': '강원', '강원도': '강원', '충청북도': '충북',
    '충청남도': '충남', '전북특별자치도': '전북', '전라북도': '전북', '전라남도': '전남',
    '경상북도': '경북', '경상남도': '경남', '제주특별자치도': '제주', '제주도': '제주',
}
TARGET_BTYPES = ['관광호텔', '일반호텔', '숙박업(생활)', '휴양콘도미니엄업']

def region_of(addr):
    for k, v in SIDO.items():
        if addr.startswith(k):
            return v
    return ''

def district_of(addr):
    m = re.match(r'\S+\s+(\S+?[구군시])\s', addr + ' ')
    return m.group(1) if m else ''

def s(v):
    return '' if v is None else str(v).strip()

def num(v):
    try:
        n = int(float(v))
        return n if n >= 0 else 0
    except (TypeError, ValueError):
        return 0

# ── 1. Excel → 관리번호 dedup 원장 ─────────────────────────
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
seen, dups = {}, 0
for sn in wb.sheetnames:
    for r in wb[sn].iter_rows(min_row=2, values_only=True):
        if not r or (r[4] is None and r[24] is None):
            continue
        mn = s(r[4]) or ('NOKEY:' + s(r[24]) + '|' + s(r[21]))
        if mn in seen:
            dups += 1
            continue
        lot, road = s(r[21]), s(r[22])
        seen[mn] = {
            'mn': mn, 'name': s(r[24]), 'operator': s(r[5]),
            'contract_type': s(r[6]), 'contract_period': s(r[7]),
            'license_date': s(r[8]), 'status': s(r[11]), 'closed_date': s(r[14]),
            'phone': s(r[18]), 'lot_address': lot, 'road_address': road,
            'btype': s(r[28]), 'rk': num(r[38]), 'rw': num(r[39]),
            'owner_type': s(r[43]), 'last_modified': s(r[25]),
            'region': region_of(road) or region_of(lot),
            'district': district_of(road or lot), 'sheet': sn,
        }
        seen[mn]['rt'] = seen[mn]['rk'] + seen[mn]['rw']
wb.close()
REG = list(seen.values())
print(f'원장 고유 사업장 {len(REG)}건 (시트 간 중복 {dups}건 제거)')

# ── 2. Dataset A 추출 (index.html 의 HOTELS) ───────────────
html = io.open(HTML, encoding='utf-8').read()
import subprocess, tempfile
node_extract = r'''
const fs=require('fs');
const L=fs.readFileSync(process.argv[2],'utf8').split('\n');
const HOTELS=eval(L.find(x=>x.startsWith('const HOTELS = ')).replace(/^const HOTELS = /,'').replace(/;$/,''));
const EX_AREA=eval(L.find(x=>x.startsWith('const EX_AREA')).replace(/^const EX_AREA = /,'').replace(/;$/,''));
const i0=L.findIndex(x=>x.trim().startsWith('const JOBS = ['));let i1=i0;while(L[i1].trim()!=='];')i1++;
const JOBS=eval(L.slice(i0,i1+1).join('\n').replace(/const JOBS = /,''));
fs.writeFileSync(process.argv[3], JSON.stringify({
  hotels: HOTELS.map(h=>({name:h[0],star:h[1],region:EX_AREA[h[2]],rooms:h[3]})),
  jobHotels: [...new Set(JOBS.map(j=>j.hotel))],
}));
'''
tf = os.path.join(BASE, '_extract.js')
io.open(tf, 'w', encoding='utf-8').write(node_extract)
out = os.path.join(BASE, '_a.json')
subprocess.run(['node', tf, HTML, out], check=True)
A_DATA = json.load(io.open(out, encoding='utf-8'))
os.remove(tf); os.remove(out)
A = A_DATA['hotels']
print(f'Dataset A: {len(A)}개 등급유효 호텔')

# ── 3. 이름 정규화 + 수기 별칭 ─────────────────────────────
def norm(n):
    n = re.sub(r'[（(]주[)）]|㈜|주식회사|\(유\)|유한회사', '', str(n or ''))
    n = re.sub(r'[()（）\[\]【】\'\"·.,&\-–—ⅠⅡⅢ\s]', '', n)
    n = re.sub(r'HOTEL', '호텔', n, flags=re.I)
    return n.upper()

def core(n):
    c = re.sub(r'관광호텔|호텔|리조트|레지던스', '', norm(n))
    return c if len(c) >= 3 else norm(n)

# 자동매칭이 놓치는 상호(법인명↔브랜드명) — 전부 원문 대조로 확인한 것만 등록
ALIAS_A = {   # Dataset A 호텔명 → 원장 사업장명
    '웨스틴 조선 부산': '(주)조선호텔앤리조트 부산',
    '호텔노아': '주식회사 경원건설 호텔노아',
}
ALIAS_JOB = {  # 구인공고 호텔명 → 원장 사업장명
    '씨클라우드 호텔 해운대': '코오롱씨클라우드호텔',
    '웨스틴조선 부산': '(주)조선호텔앤리조트 부산',
    '시그니엘 부산 (호텔롯데)': '(주)호텔롯데 시그니엘 부산',
    'L7 해운대 (호텔롯데)': '호텔롯데 L7해운대',
    '그랩디오션 송도호텔': '그랩 디 오션 송도',
    '빌라쥬 드 아난티': '아난티 앳 부산 빌라쥬',
    '파라다이스호텔 부산': '㈜파라다이스호텔부산',
    '케니스테이 기장': '주식회사 케니스테이부산기장',
    '하단 레이어스호텔': '레이어스호텔',
    '호텔 게스후 광안리': '게스후',
    '영무파라드호텔 해운대': '영무파라드호텔',
    '호텔농심': '호텔농심',
    '라마다앙코르 부산역호텔': '라마다앙코르부산역호텔',
    '부산롯데호텔': '(주)부산롯데호텔',
    '아스티호텔 부산': '아스티호텔',
    '소노문 해운대': '소노문 해운대',
    '아난티코브': '아난티펜트하우스해운대',
}

for r in REG:
    r['n'] = norm(r['name'])
    r['c'] = core(r['name'])
by_n = {}
for r in REG:
    by_n.setdefault(r['n'], []).append(r)

# ── 4. Cross Validation Engine ─────────────────────────────
used = set()
xm = []       # A 인덱스 정렬: [status, level, b_rt, b_btype, b_name, b_mn]
xsum = {'MATCHED': 0, 'REVIEW_REQUIRED': 0, 'A_ONLY': 0, 'CONFLICT': 0}
for a in A:
    an, ac = norm(a['name']), core(a['name'])
    match, level, review = None, 0, False
    if a['name'] in ALIAS_A:
        tgt = norm(ALIAS_A[a['name']])
        cands = [b for b in by_n.get(tgt, []) if b['region'] == a['region']]
        if cands:
            match, level = cands[0], 1
    if not match:
        cands = [b for b in by_n.get(an, []) if b['region'] == a['region']]
        if cands:
            match, level = cands[0], 1
    if not match:
        cands = by_n.get(an, [])
        if cands:
            match, level, review = cands[0], 2, True
    if not match:
        cands = [b for b in REG if b['region'] == a['region'] and b['mn'] not in used and
                 ((len(an) >= 5 and an in b['n']) or (len(b['n']) >= 5 and b['n'] in an))]
        if cands:
            cands.sort(key=lambda b: abs(b['rt'] - a['rooms']))
            match, level = cands[0], 3
            review = len(cands) > 2
    if not match and len(ac) >= 4:
        cands = [b for b in REG if b['region'] == a['region'] and b['mn'] not in used and b['c'] == ac]
        if cands:
            match, level = cands[0], 4
            review = len(cands) > 1
    if match:
        used.add(match['mn'])
        st = 'REVIEW_REQUIRED' if review else 'MATCHED'
        xsum[st] += 1
        conflict = match['rt'] > 0 and a['rooms'] > 0 and match['rt'] != a['rooms']
        if conflict and st == 'MATCHED':
            xsum['CONFLICT'] += 1
        xm.append([1 if st == 'MATCHED' else 2, level, match['rt'], match['btype'], match['name'], match['mn']])
        match['a_name'] = a['name']; match['a_rooms'] = a['rooms']; match['a_star'] = a['star']
    else:
        xsum['A_ONLY'] += 1
        xm.append([0, 0, 0, '', '', ''])
xsum['B_ONLY'] = len(REG) - len(used)
print('교차검증:', json.dumps(xsum, ensure_ascii=False))

# ── 5. 구인공고 ↔ 원장 연결 ───────────────────────────────
job_mns = set()
job_map = {}
busan = [r for r in REG if r['region'] == '부산']
for h in A_DATA['jobHotels']:
    tgt = None
    if h in ALIAS_JOB:
        tn = norm(ALIAS_JOB[h])
        tgt = next((r for r in busan if r['n'] == tn), None)
    if not tgt:
        hn = norm(h)
        tgt = next((r for r in busan if r['n'] == hn or
                    (len(hn) >= 5 and hn in r['n']) or (len(r['n']) >= 5 and r['n'] in hn)), None)
    if tgt:
        job_mns.add(tgt['mn'])
        job_map[h] = tgt['mn']
print(f'구인공고 호텔 {len(A_DATA["jobHotels"])}종 중 원장 연결 {len(job_map)}종 / 시설 {len(job_mns)}곳')


# ---- 6.5 운영·브랜드 구분 (전국 호텔 현황 탭용) ----
#  원장에는 '분양형' 필드가 없다(건물소유구분명은 자가/임대뿐) ->
#  분양 여부는 단정하지 않고, 사업장명 브랜드 매칭 + 업태로만 구분한다.
BRAND_RULES = [
    (r'조선호텔|그랜드조선|웨스틴조선|조선팰리스|그래비티', '조선', 'K'),
    (r'메리어트|Marriott|페어필드|코트야드|목시|MOXY|웨스틴|쉐라톤|포포인츠|메리디앙', '메리어트 계열', 'G'),
    (r'힐튼|Hilton|콘래드', '힐튼', 'G'),
    (r'하얏트|Hyatt', '하얏트', 'G'),
    (r'이비스|노보텔|머큐어|소피텔|풀만|아코르', '아코르 계열', 'G'),
    (r'홀리데이인|홀리데이 인|인터컨티넨탈|인디고', 'IHG 계열', 'G'),
    (r'라마다|Ramada|윈덤|Wyndham|하워드존슨|데이즈호텔', '윈덤 계열', 'G'),
    (r'반얀트리', '반얀트리', 'G'),
    (r'베스트웨스턴', '베스트웨스턴', 'G'),
    (r'오크우드|Oakwood', '오크우드', 'G'),
    (r'이스틴|EASTIN', '이스틴', 'G'),
    (r'롯데호텔|롯데시티|시그니엘|L7|롯데리조트', '롯데', 'K'),
    (r'신라스테이|신라호텔|신라모노그램', '신라', 'K'),

    (r'신세계', '신세계', 'K'),
    (r'켄싱턴', '켄싱턴', 'K'),
    (r'소노문|소노캄|소노벨|소노펠리체|대명리조트|대명콘도', '소노(대명)', 'K'),
    (r'아난티', '아난티', 'K'),
    (r'토요코인', '토요코인', 'K'),
    (r'글래드|GLAD', '글래드', 'K'),
    (r'스카이파크', '스카이파크', 'K'),
    (r'베니키아', '베니키아', 'K'),
    (r'라한호텔|라한셀렉트', '라한', 'K'),
    (r'브라운도트', '브라운도트', 'F'),
    (r'하운드', '하운드', 'F'),
    (r'넘버25|No[.]?25', '넘버25', 'F'),
    (r'야자|YAJA', '야자', 'F'),
    (r'어반스테이', '어반스테이', 'F'),
]
_BRAND_C = [(re.compile(p_, re.I), lb, tier) for p_, lb, tier in BRAND_RULES]

def classify(r):
    for pat, lb, tier in _BRAND_C:
        if pat.search(r['name']):
            return tier, lb
    if r['btype'] == '숙박업(생활)':
        return 'R', ''
    if r['btype'] == '휴양콘도미니엄업':
        return 'C', ''
    return 'I', ''

CATS = [
    ('G', '글로벌 체인'), ('K', '국내 체인'), ('F', '국내 프랜차이즈'),
    ('R', '생활숙박·레지던스 (비체인)'), ('C', '휴양콘도 (비체인)'), ('I', '독립 관광·일반호텔'),
]
cls_sum = {k: {'label': lb, 'cnt': 0, 'rooms': 0, 'b_cnt': 0, 'b_rooms': 0, 'brands': {}} for k, lb in CATS}
for r in REG:
    if r['btype'] not in TARGET_BTYPES:
        continue
    tier, lb = classify(r)
    c = cls_sum[tier]
    c['cnt'] += 1
    c['rooms'] += r['rt']
    if r['region'] == '부산':
        c['b_cnt'] += 1
        c['b_rooms'] += r['rt']
    if lb:
        b = c['brands'].setdefault(lb, [0, 0])
        b[0] += 1
        b[1] += r['rt']
REG_CLASS = []
for k, lb in CATS:
    c = cls_sum[k]
    top = sorted(c['brands'].items(), key=lambda x: -x[1][0])[:8]
    REG_CLASS.append({'key': k, 'label': lb, 'cnt': c['cnt'], 'rooms': c['rooms'],
                      'b_cnt': c['b_cnt'], 'b_rooms': c['b_rooms'],
                      'top': [[n_, v[0], v[1]] for n_, v in top]})
print('운영·브랜드 구분:', json.dumps([[c['label'], c['cnt'], c['rooms']] for c in REG_CLASS], ensure_ascii=False))

# ── 6. 출력 데이터 생성 ────────────────────────────────────
BT = ['관광호텔', '일반호텔', '숙박업(생활)', '휴양콘도미니엄업', '숙박업 기타', '여관업', '여인숙업']
bti = {b: i for i, b in enumerate(BT)}

summary = {'total': len(REG), 'dups_removed': dups,
           'target': sum(1 for r in REG if r['btype'] in TARGET_BTYPES),
           'target30': sum(1 for r in REG if r['btype'] in TARGET_BTYPES and r['rt'] >= 30),
           'btypes': BT, 'regions': {}}
for r in REG:
    reg = summary['regions'].setdefault(r['region'] or '?', [0] * (len(BT) + 1))
    reg[bti.get(r['btype'], len(BT) - 1)] += 1
    reg[len(BT)] += 1

# 부산 원장 (compact): [name, btypeIdx, rt, operator, contract, district, road, lic, mn, flags, a_name, a_rooms]
regb = []
for r in sorted(busan, key=lambda x: -x['rt']):
    flags = (1 if r['mn'] in used else 0) | (2 if r['mn'] in job_mns else 0)
    regb.append([r['name'], bti.get(r['btype'], 6), r['rt'], r['operator'],
                 r['contract_type'], r['district'], r['road_address'][:46], r['license_date'],
                 r['mn'], flags, r.get('a_name', ''), r.get('a_rooms', 0)])

json.dump(regb, io.open(os.path.join(BASE, 'data', 'registry_busan.json'), 'w', encoding='utf-8'), ensure_ascii=False)
json.dump(summary, io.open(os.path.join(BASE, 'data', 'registry_summary.json'), 'w', encoding='utf-8'), ensure_ascii=False)
json.dump({'xm': xm, 'sum': xsum}, io.open(os.path.join(BASE, 'data', 'xmatch.json'), 'w', encoding='utf-8'), ensure_ascii=False)

# ---- 6.6 탐색기 확장 (REGX) : 등급DB에 없는 원장 타깃시설 ----
#  전국은 30실 이상 또는 브랜드 매칭 시설만, 부산은 타깃업태 전부.
#  A와 매칭된 시설은 탐색기에 이미 있으므로 제외(중복 방지).
mn2job = {}
for jh, mn_ in job_map.items():
    mn2job.setdefault(mn_, jh)
bti_x = {'관광호텔': 0, '일반호텔': 1, '숙박업(생활)': 2, '휴양콘도미니엄업': 3}
REGX = []
for r in REG:
    if r['btype'] not in TARGET_BTYPES or r['mn'] in used:
        continue
    tier_x, lb_x = classify(r)
    # 100객실 이하 제외 (2026-08-31 지시 — 소형 시설은 도급 영업 의미 없음)
    if r['rt'] <= 100:
        continue
    REGX.append([r['name'], r['region'], r['district'], bti_x[r['btype']], r['rt'],
                 mn2job.get(r['mn'], ''), r['operator'] if r['operator'] not in ('', '-') else '',
                 r['owner_type']])
REGX.sort(key=lambda x: -x[4])
print(f'탐색기 확장(REGX): {len(REGX)}개 시설 (부산 {sum(1 for x in REGX if x[1] == "부산")})')


# ── 7. index.html 데이터 블록 주입 ─────────────────────────
block = ('/*==REGISTRY_DATA_START==*/\n'
         '/* Dataset B — 숙박업 인허가 원장 (숙박업 리스트_260220.xlsx · registry_build.py 생성).\n'
         '   Dataset A(등급유효 835)와 모집단이 다르므로 서로 덮어쓰지 않는다. */\n'
         f'const REG_SUMMARY = {json.dumps(summary, ensure_ascii=False)};\n'
         f'const REGB = {json.dumps(regb, ensure_ascii=False)};\n'
         f'const XM = {json.dumps(xm, ensure_ascii=False)};\n'
         f'const XM_SUM = {json.dumps(xsum, ensure_ascii=False)};\n'
         f'const JOB_REG_MAP = {json.dumps(job_map, ensure_ascii=False)};\n'
         f'const REG_CLASS = {json.dumps(REG_CLASS, ensure_ascii=False)};\n'
         f'const REGX = {json.dumps(REGX, ensure_ascii=False)};\n'
         '/*==REGISTRY_DATA_END==*/')
pat = re.compile(r'/\*==REGISTRY_DATA_START==\*/.*?/\*==REGISTRY_DATA_END==\*/', re.S)
if pat.search(html):
    html = pat.sub(lambda m: block, html)
    print('index.html 데이터 블록 갱신')
else:
    print('경고: index.html 에 REGISTRY_DATA 마커 없음 — UI 패치를 먼저 적용할 것')
io.open(HTML, 'w', encoding='utf-8').write(html)
print('완료')
